import requests
import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup
import os
from re import search

Headers = ["ID", "Specification", "Year","Make", "Model", "Price ($)", "Odometer (km)","Transmission", "Engine", "Body", "Location", "Seller", "URL"]
class carInfo:

	#set values
	def setTitle(self, title):
		self.title = title.replace(str(self.year), "").replace(self.make, "").replace(self.model,"")

	def setYear(self, year):
		self.year = int(year)

	def setMake(self, make):
		self.make = make

	def setModel(self, model):
		self.model = model

	def setPrice(self, price):
		self.price = int(price)

	def setOdometer(self, odometer):
		self.odometer = int(odometer)

	def setTransmission(self, transmission):
		self.transmission = transmission

	def setBodyStyle(self, bodyStyle):
		self.bodyStyle = bodyStyle

	def setEngine(self, engine):
		self.engine = engine

	def setLocation(self, location):
		self.location = location

	def setSeller(self, seller):
		self.seller = seller

	def setURL(self, url):
		self.url = 'https://www.carsales.com.au' + url

	def setID(self, ID):
		self.ID = ID

	# return values
	def getTitle(self):
		return self.title

	def getYear(self):
		return self.year

	def getMake(self):
		return self.make

	def getModel(self):
		return self.model

	def getPrice(self):
		return self.price 

	def getOdometer(self):
		return self.odometer

	def getTransmission(self):
		return self.transmission 

	def getBodyStyle(self):
		return self.bodyStyle 

	def getEngine(self):
		return self.engine

	def getLocation(self):
		return self.location 

	def getSeller(self):
		return self.seller

	def getURL(self):
		return self.url

	def getID(self):
		return self.ID

def getFiles():

	path = 'CarSaleFiles/'
	osFiles = os.listdir(path)
	
	files = []
	for index, file in enumerate(osFiles):
		if file.endswith('.html'):
			files.append(file)

	return files

def getCarDivs(soup):
	''' Gets car divs '''

	cars = []
	itemlist = ["listing-item card topspot", "listing-item card showcase", "listing-item card standard"]

	for item in itemlist:
			items = soup.find_all('div', {'class': item})
			for car in items:
				cars.append(car)
	return cars

def setKey_Details(car, car_class):

	car = BeautifulSoup(str(car), 'html5lib')
	terms = ['Odometer', 'Body Style', 'Transmission', 'Engine', 'id']

	#if value for odometer exists, then concert odometer to int value
	try:
		odometer = car.find('li', {'data-type': terms[0]}).text
		odometer = odometer.split(" ")[0]
		odometer = odometer.replace(",", "")
	except AttributeError:
		odometer = 0

	car_class.setOdometer(odometer)
	car_class.setBodyStyle(car.find('li', {'data-type': terms[1]}).text)
	car_class.setTransmission(car.find('li', {'data-type': terms[2]}).text)
	car_class.setEngine(car.find('li', {'data-type': terms[3]}).text)

def setOther_info(car, car_class):
	SearchTerms = ["data-webm-make","data-webm-model","data-webm-price", 'data-webm-state', 'data-webm-vehcategory', 'id']

	#splits div to get first one
	info = ((str(car).split('\n')[0]).split(" "))
	# gets rid of the fisrt 4
	del info[0:4]
	
	# checks to see if there is part of a name on its own e.g "Model", "3" should be "Model 3"
	idToDel = []
	for i in range(0, len(info)):
		if 'data' not in info[i] and 'id' not in info[i]:
			info[i-1] = '{} {}'.format(info[i-1], info[i])
			idToDel.append(i)

	#Deletes from info lone item
	for i in idToDel:
		del info[i]

	for i in range(0, len(info)):
		info[i] = info[i].split("=")

	#convert 2D array to dict
	carDict = {  v[0]:v[1] for k,v in enumerate(info)}

	car_class.setMake(carDict[SearchTerms[0]].replace('"', ""))
	car_class.setModel(carDict[SearchTerms[1]].replace('"', ""))
	car_class.setPrice(carDict[SearchTerms[2]].replace('"', ""))
	car_class.setLocation(carDict[SearchTerms[3]].replace('"', ""))
	car_class.setSeller(carDict[SearchTerms[4]].replace('"', ""))
	car_class.setID(carDict[SearchTerms[5]].replace('"', ""))

def setURL_Year_Title(car, car_class):

	car = BeautifulSoup(str(car), 'html5lib')

	title = car.find('div', {'class': 'card-body'})
	url = title.find('a',{'class':'js-encode-search'})['href']
	year = url.split('-')[0].split('/')[-1]

	car_class.setURL(url)
	car_class.setYear(year)
	car_class.setTitle(title.find('a').text)

def writeToFile(CarsArray):

	wb = openpyxl.load_workbook(filename = 'cars.xlsx')
	ws = wb['Raw_Data']

	#set up Headers
	for col, header in enumerate(Headers):
		ws.cell(column=col + 1, row=1, value=header)

	for row, car in enumerate(CarsArray):

		ws.cell(row=row+2, column=1, value=car.getID())
		ws.cell(row=row+2, column=2, value=car.getTitle())
		ws.cell(row=row+2, column=3, value=car.getYear())
		ws.cell(row=row+2, column=4, value=car.getMake())
		ws.cell(row=row+2, column=5, value=car.getModel())
		ws.cell(row=row+2, column=6, value=car.getPrice())
		ws.cell(row=row+2, column=7, value=car.getOdometer())
		ws.cell(row=row+2, column=8, value=car.getTransmission())
		ws.cell(row=row+2, column=9, value=car.getEngine())
		ws.cell(row=row+2, column=10, value=car.getBodyStyle())
		ws.cell(row=row+2, column=11, value=car.getLocation())
		ws.cell(row=row+2, column=12, value=car.getSeller())
		ws.cell(row=row+2, column=13, value=car.getURL()).hyperlink = (car.getURL())

	wb.save('cars.xlsx')
def main():

	CarsArray = []

	files = getFiles()

	for file in files:
		print(file)
		page = open("CarSaleFiles/" + file)
		soup = BeautifulSoup(page.read(), 'html5lib')
		cars = getCarDivs(soup)

		for car in cars:

			car_class = carInfo()

			#detail split into 3 becasue of the way the website works 
			setKey_Details(car, car_class)
			setOther_info(car, car_class)
			setURL_Year_Title(car, car_class)

			CarsArray.append(car_class)
			
	writeToFile(CarsArray)
if __name__ == "__main__":

	main()





