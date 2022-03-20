from tkinter import OFF
import requests
from bs4 import BeautifulSoup
import json
import openpyxl
from openpyxl.styles import Font
import os

BASE_URL="https://www.carsales.com.au"

headers={
    'authority':'www.carsales.com.au',
    'upgrade-insecure-requests':'1',
    'user-agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36',
    'accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'sec-gpc':'1',
    'sec-fetch-site':'same-origin',
    'sec-fetch-mode':'navigate',
    'sec-fetch-user':'?1',
    'sec-fetch-dest':'document',
    'referer':'https://www.carsales.com.au/',
    'accept-language':'en-GB,en-US;q=0.9,en;q=0.8',
}


def writeToExcel(cars):
    column_headers = ["ID","Year","Make", "Model", "Price ($)", "Odometer (km)","Transmission", "Body", "Engine", "Seller","Location", "URL"]

    wb = openpyxl.load_workbook(filename = './cars.xlsx')
    ws = None
    try:
        ws = wb['Raw_Data']

    except KeyError:
        sheet = wb['Sheet1']
        sheet.title = 'Raw_Data'
        ws = wb['Raw_Data']

    # Write column headers
    for col, header in enumerate(column_headers):
        c = ws.cell(column=col + 1, row=1)
        c.font = Font(bold=True)
        c.value = header

    #For every car in car array
    for row, car in enumerate(cars):
        #For every key in car dict
        for index, value in enumerate(car.items()):
            #row+1 is row offset, and index+1 is column offset
            c = ws.cell(row=row+2, column=index+1)
            
            if value[0] == 'price':
                c.number_format = '$#,###'
            elif value[0] == 'odometer':
                c.number_format = '#,###'
            elif value[0] == 'link':
                c.style = 'Hyperlink'

            c.value=value[1]

    wb.save('cars.xlsx')

def getCars(soup, cars):

    container = soup.find('div',{'class':'listing-items'})

    items = container.find_all('div',{'class':'listing-item card showcase'})
    for item in items:
        cars.append(getCarDetails(item))
    
    items = container.find_all('div',{'class':'listing-item card topspot'})
    for item in items:
        cars.append(getCarDetails(item))

    items = container.find_all('div',{'class':'listing-item card standard'})
    for item in items:
        cars.append(getCarDetails(item))

    nextPage = soup.find('a',{'class':'page-link next'})

    if nextPage != None:
        nextPage = nextPage.get('href')
        print(f'requesting next page {BASE_URL}{nextPage}')
        r = requests.get(f"{BASE_URL}{nextPage}", headers=headers)
        soup = BeautifulSoup(r.content, 'html.parser')
        cars = getCars(soup, cars)

    return cars

def processTitle(title):
    seperator = ' '
    titleSplit = title.split(seperator)

    return [titleSplit[0], titleSplit[1], seperator.join(titleSplit[2:])]

def cleanOdometer(odomter):

    odomter = odomter.replace(',','')
    odomter = int(odomter.strip(' km'))

    return odomter

def cleanPrice(price):

    price = price.strip('$')
    price = price.strip('*')
    price = int(price.replace(',',''))
    
    return price    

def getCarDetails(item):

    car_id = item.get('id')
    titleRaw = item.find('a',{'data-webm-clickvalue':'sv-title'}).text
    odometerRaw = item.find('li',{'data-type':'Odometer'}).text
    transmission = item.find('li',{'data-type':'Transmission'}).text
    engine = item.find('li',{'data-type':'Engine'}).text
    priceRaw = item.find('a',{'data-webm-clickvalue':'sv-price'}).text
    body = item.find('li',{'data-type':'Body Style'}).text
    seller = item.find('div',{'class':'seller-type'}).text.split(' ')[0]
    location = item.find('div',{'class':'seller-location d-flex'}).text
    link = f"https://www.carsales.com.au{item.find('a',{'data-webm-clickvalue':'sv-title'}).get('href')}"

    year, brand, model = processTitle(titleRaw)
    car = {
        "id":car_id,
        "year": int(year),
        "make":brand,
        "model":model,
        "price": cleanPrice(priceRaw),
        "odometer": cleanOdometer(odometerRaw),
        "transmission": transmission,
        "body":body,
        "engine":engine,
        "seller":seller,
        "location":location,
        "link": link
    }
    
    return car

def main():

    url = open("LINK_HERE.txt",'r').read()
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.content, 'html.parser')
    cars = getCars(soup, [])

    writeToExcel(cars)

if __name__ == "__main__":
    main()